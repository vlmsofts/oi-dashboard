"""
build_prelim_oi.py — VLM ICE Preliminary Open Interest Monitor

Reads the ICE "Preliminary Open Interest - Futures" CSV (manually downloaded from
ice.com/report/114 — the report is reCAPTCHA-gated, so there is no unattended fetch)
and compares it against OFFICIAL open interest history in oi_data.csv.

The prelim is an EARLY look at a session whose official number does not exist yet:
a report stamped 2026-08-14 describes the 08-14 session, which Bloomberg won't
publish until master fetch runs at 09:30. The baseline is therefore the PRIOR
session (T-1), which is already on disk — so DoD/WoW/MoM are fully computable at
5am and produce identical numbers at 10am. Nothing is stale, nothing shifts.

Every current value is PRELIMINARY; every baseline is OFFICIAL. That mixed basis
is stated on the output rather than left implicit.

Usage:
  python build_prelim_oi.py                      # newest CSV in Downloads
  python build_prelim_oi.py --csv path/to.csv    # explicit file
  python build_prelim_oi.py --no-png             # xlsx only

Requires: playwright (PNG), openpyxl (xlsx)
"""

import csv, argparse, pathlib, re, sys
from datetime import datetime, date

BASE_DIR = pathlib.Path(__file__).parent
OI_FILE  = BASE_DIR / 'data' / 'oi_data.csv'
OUT_DIR  = BASE_DIR / 'output' / 'prelim'
DOWNLOADS = pathlib.Path.home() / 'Downloads'

# Sugar No.11 only. The same ICE file also carries SF = SUGAR 16, a different
# contract entirely -- matching on the name would silently pull both, so the
# filter is pinned to the exact symbol and the name is asserted below.
COMMODITIES = ['CT', 'KC', 'CC', 'SB']
# PNG card tiling only -- see build_html(). Data/xlsx keep COMMODITIES order.
GRID_ORDER  = ['CT', 'CC', 'SB', 'KC']
COMM_NAMES  = {'CT': 'COTTON', 'KC': 'COFFEE', 'CC': 'COCOA', 'SB': 'SUGAR No.11'}
NAME_GUARD  = {'CT': 'COTTON', 'KC': 'COFFEE', 'CC': 'COCOA', 'SB': 'SUGAR 11'}

# Baseline offsets in TRADING SESSIONS present in oi_data.csv -- not calendar
# days, so holidays never silently skew a window.
WINDOWS = [('DoD', 1), ('WoW', 5), ('MoM', 21)]

MONTHS = {1:'JAN',2:'FEB',3:'MAR',4:'APR',5:'MAY',6:'JUN',
          7:'JUL',8:'AUG',9:'SEP',10:'OCT',11:'NOV',12:'DEC'}


# ── Contract mapping ─────────────────────────────────────────────────────────
def contract_key(commodity, first_notice, last_trade):
    """Map an official row to the ICE prelim's 'Contract Month' label (e.g. DEC26).

    The anchor field is commodity-specific and this is NOT interchangeable:
      * CT/CC/KC -- last_trade falls IN the delivery month; FND is the month before.
      * SB       -- cash-settle pattern, LTD falls in the month BEFORE delivery
                    (SBOCT1 expires 09-30), so FND is the correct anchor.
    Verified: this split maps 36/36 official contracts, and the resulting DoD
    reconciles exactly against ICE's own change column.
    """
    anchor = first_notice if commodity == 'SB' else last_trade
    if not anchor:
        return None
    try:
        d = datetime.strptime(anchor, '%Y-%m-%d')
    except ValueError:
        return None
    return f'{MONTHS[d.month]}{d.strftime("%y")}'


def find_prelim_csv(explicit=None):
    """Newest ICE prelim CSV by Report Date (NOT by filename or mtime -- the
    download name is an arbitrary '(n)' counter and mtime reflects when it was
    saved, neither of which identifies the session)."""
    if explicit:
        p = pathlib.Path(explicit)
        if not p.exists():
            sys.exit(f'ERROR: {p} not found')
        return p
    cands = list(DOWNLOADS.glob('PreliminaryOpenInterestFutures*.csv'))
    if not cands:
        sys.exit(f'ERROR: no PreliminaryOpenInterestFutures*.csv in {DOWNLOADS}')
    best, best_date = None, ''
    for c in cands:
        try:
            with c.open(encoding='utf-8-sig') as f:
                row = next(csv.DictReader(f), None)
            d = (row or {}).get('Report Date', '')
            if d > best_date:
                best, best_date = c, d
        except Exception:
            continue
    if not best:
        sys.exit('ERROR: could not read a Report Date from any candidate CSV')
    return best


def load_prelim(path):
    """Return (report_date, {commodity: {contract_month: oi}}, {(c,m): ice_chg})."""
    rows = list(csv.DictReader(path.open(encoding='utf-8-sig')))
    if not rows:
        sys.exit(f'ERROR: {path} is empty')
    report_date = rows[0]['Report Date']
    if len({r['Report Date'] for r in rows}) > 1:
        sys.exit('ERROR: prelim CSV contains multiple Report Dates')

    data, ice_chg = {c: {} for c in COMMODITIES}, {}
    for r in rows:
        sym = r['Commodity Symbol'].strip().upper()
        if sym not in COMMODITIES:
            continue
        name = r['Commodity Name'].upper()
        if NAME_GUARD[sym] not in name:
            sys.exit(f'ERROR: symbol {sym} carried unexpected name {name!r} — '
                     f'expected {NAME_GUARD[sym]!r}. Refusing to guess.')
        month = r['Contract Month'].strip().upper()
        try:
            data[sym][month] = int(r['Total Preliminary Open Interest'])
            ice_chg[(sym, month)] = int(r['Open Interest Change'])
        except (ValueError, TypeError):
            continue
    return report_date, data, ice_chg


def load_official():
    """Return (sessions_sorted, {date: {(commodity, contract_month): row}})."""
    rows = list(csv.DictReader(OI_FILE.open(encoding='utf-8')))
    by_date, sessions = {}, set()
    for r in rows:
        c = r['commodity']
        if c not in COMMODITIES:
            continue
        key = contract_key(c, r.get('first_notice',''), r.get('last_trade',''))
        if not key:
            continue
        d = r['date']
        sessions.add(d)
        slot = by_date.setdefault(d, {})
        prior = slot.get((c, key))
        if prior is not None and prior['contract'] != r['contract']:
            # Roll-boundary collision: on the day a generic rolls, two slots can
            # briefly share a last_trade/first_notice and map to the same month
            # (observed 2026-05-13, CCMAY1 + CCMAY2 both -> MAY27). Last-write-wins
            # would silently substitute one contract's OI for the other's, so keep
            # the LOWER generation -- the nearer generic is the live contract the
            # prelim's month label refers to.
            if _generation(prior['contract']) <= _generation(r['contract']):
                continue
        slot[(c, key)] = r
    return sorted(sessions), by_date


def _generation(contract):
    """Trailing slot digit of a month-specific generic (CCMAY1 -> 1)."""
    m = re.search(r'(\d+)$', contract or '')
    return int(m.group(1)) if m else 99


def pick_baselines(sessions, report_date):
    """Sessions strictly BEFORE the prelim's own session, walked back N steps.

    The prelim session itself is excluded even if official data for it happens
    to exist (i.e. when run after master fetch) -- comparing prelim against the
    official value of the SAME session would report a revision, not a flow.
    """
    prior = [s for s in sessions if s < report_date]
    out = {}
    for label, n in WINDOWS:
        out[label] = prior[-n] if len(prior) >= n else None
    return out


def build_rows(prelim, ice_chg, by_date, baselines):
    """Per-commodity: contract rows + totals. Missing baseline -> None (dash)."""
    result = {}
    for comm in COMMODITIES:
        months = prelim.get(comm, {})
        if not months:
            continue

        def sort_key(m):
            """Chronological by delivery, from the MONyy label itself."""
            mon, yr = m[:3], m[3:]
            idx = next((k for k, v in MONTHS.items() if v == mon), 99)
            return (int(yr), idx)

        rows, totals = [], {lab: 0 for lab, _ in WINDOWS}
        tot_oi, tot_base_ok = 0, {lab: True for lab, _ in WINDOWS}
        for m in sorted(months, key=sort_key):
            oi = months[m]
            tot_oi += oi
            changes = {}
            for lab, _ in WINDOWS:
                bdate = baselines.get(lab)
                brow = by_date.get(bdate, {}).get((comm, m)) if bdate else None
                if brow and brow.get('open_int'):
                    chg = oi - int(brow['open_int'])
                    changes[lab] = chg
                    totals[lab] += chg
                else:
                    # No official baseline: contract not carried by Bloomberg's
                    # generic chain (deep back months) or not yet listed then.
                    changes[lab] = None
                    tot_base_ok[lab] = False
            rows.append({'month': m, 'oi': oi, 'chg': changes,
                         'ice_chg': ice_chg.get((comm, m))})
        # Totals sum every contract that HAS an official baseline, and carry a
        # partial flag when one or more were excluded. Nulling the whole total
        # over a 2-lot back month would blank the most-watched number on the page.
        result[comm] = {
            'rows': rows,
            'total_oi': tot_oi,
            'total_chg': {lab: totals[lab] for lab, _ in WINDOWS},
            'partial': {lab: not tot_base_ok[lab] for lab, _ in WINDOWS},
            'excluded': {lab: sum(1 for r in rows if r['chg'][lab] is None)
                         for lab, _ in WINDOWS},
        }
    return result


# ── Rendering ────────────────────────────────────────────────────────────────
# VLM PNG Master Palette (see 'VLM PNG Master pallate.md') -- these exact hexes
# are the house standard; do not substitute near-variants.
NAVY   = '#1a1a2e'   # header bg, primary data text
GOLD   = '#c9a227'   # header rule, section banners
WHITE  = '#ffffff'   # body bg, even rows
ALT    = '#f9f9f9'   # odd rows, footer
BORDER = '#e5e7eb'
DARK   = '#222222'   # row labels
GRAY   = '#444444'   # secondary body text
LGRAY  = '#888888'   # small-caps labels, neutral/zero
MUTED  = '#aaaaaa'
DKROW  = '#2c3e50'   # totals row
GREEN  = '#15803d'
RED    = '#c0392b'
AMBER  = '#EF9F27'   # key highlight values

# OI values sit on white rows, so they use NAVY (the palette's "open interest"
# column colour) rather than gold -- gold on light is the low-contrast case.
OI_COL     = '#1a2535'
OI_COL_TOT = '#c8d8e8'   # open interest on the dark totals row

# GREEN/RED are tuned for contrast on WHITE rows. The totals row sits on DKROW
# (#2c3e50, a dark slate) where those same hexes go muddy -- brighten both for
# that row only, same reasoning as OI_COL_TOT above.
GREEN_TOT = '#4ade80'
RED_TOT   = '#f87171'
LGRAY_TOT = '#c8d0d8'


def fc(v):
    if v is None:
        return '—'
    return f'+{v:,}' if v > 0 else f'{v:,}'


def cc(v, on_dark=False):
    if v is None:
        return LGRAY_TOT if on_dark else LGRAY
    if on_dark:
        return GREEN_TOT if v > 0 else RED_TOT if v < 0 else LGRAY_TOT
    return GREEN if v > 0 else RED if v < 0 else LGRAY


def build_html(report_date, built, baselines, src_name):
    # ICE's own change column is deliberately NOT shown: it is identical to the
    # computed DoD on every row (that equality is the mapping's self-check), so
    # displaying both just prints the same number twice. The check still runs and
    # is reported in the build log.
    COLS = '84px 116px 104px 104px 104px'
    blocks = ''
    # Card order is a LAYOUT choice, not the canonical CT/KC/CC/SB order used
    # everywhere else: CT(11 rows) pairs with CC(10) on top so their heights
    # match, leaving SB(12)/KC(8) to absorb the ragged edge along the bottom.
    for comm in GRID_ORDER:
        d = built.get(comm)
        if not d:
            continue
        body = ''
        for i, r in enumerate(d['rows']):
            bg = WHITE if i % 2 == 0 else ALT
            body += f"""
        <div style="display:grid;grid-template-columns:{COLS};background:{bg};
                    padding:6px 16px;align-items:center;border-bottom:1px solid {BORDER};">
          <div style="font-size:15px;font-weight:600;color:{DARK};">{r['month']}</div>
          <div style="font-size:16px;font-weight:700;color:{OI_COL};text-align:right;">{r['oi']:,}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(r['chg']['DoD'])};text-align:right;">{fc(r['chg']['DoD'])}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(r['chg']['WoW'])};text-align:right;">{fc(r['chg']['WoW'])}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(r['chg']['MoM'])};text-align:right;">{fc(r['chg']['MoM'])}</div>
        </div>"""
        t = d['total_chg']
        star = lambda lab: '*' if d['partial'][lab] else ''
        body += f"""
        <div style="display:grid;grid-template-columns:{COLS};background:{DKROW};
                    padding:9px 16px;align-items:center;">
          <div style="font-size:14px;font-weight:700;color:{GOLD};letter-spacing:1px;">TOTAL</div>
          <div style="font-size:16px;font-weight:700;color:{OI_COL_TOT};text-align:right;">{d['total_oi']:,}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(t['DoD'], True)};text-align:right;">{fc(t['DoD'])}{star('DoD')}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(t['WoW'], True)};text-align:right;">{fc(t['WoW'])}{star('WoW')}</div>
          <div style="font-size:15px;font-weight:700;color:{cc(t['MoM'], True)};text-align:right;">{fc(t['MoM'])}{star('MoM')}</div>
        </div>"""

        # Each commodity is a self-contained card; the cards tile 2-up so the
        # page stays short enough to render without clipping and reads at a
        # glance on a phone.
        blocks += f"""
<div style="border:1px solid {BORDER};">
  <div style="background:{GOLD};border-left:6px solid {NAVY};padding:9px 18px;">
    <span style="font-size:14px;font-weight:700;letter-spacing:2px;color:{NAVY};text-transform:uppercase;">{COMM_NAMES[comm]}</span>
  </div>
  <div style="display:grid;grid-template-columns:{COLS};background:{ALT};
              padding:8px 16px;border-bottom:2px solid {BORDER};">
    {''.join(f'<div style="font-size:13px;font-weight:600;color:{LGRAY};text-align:{"left" if h=="MONTH" else "right"};letter-spacing:1px;text-transform:uppercase;">{h}</div>' for h in ['MONTH','PRELIM OI','DoD','WoW','MoM'])}
  </div>{body}
</div>"""

    b = baselines
    note = (f'<div style="font-size:13px;color:{LGRAY};padding:10px 20px;">'
            '* change total sums only contracts with an official baseline; '
            'deep back months not carried by the generic chain are excluded from '
            'the change columns (their OI is still included in PRELIM OI).</div>')

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>* {{ box-sizing:border-box; margin:0; padding:0; }}
body {{ background:{WHITE}; font-family:Arial,sans-serif; width:1120px; }}</style>
</head><body>

<div style="background:{NAVY};padding:28px 36px 22px;border-bottom:4px solid {GOLD};
            display:flex;justify-content:space-between;align-items:flex-start;">
  <div>
    <div style="font-size:32px;font-weight:700;color:{WHITE};line-height:1;">VLM Commodities &mdash; Softs Desk</div>
    <div style="font-size:16px;font-weight:600;color:{GOLD};letter-spacing:1px;
                text-transform:uppercase;margin-top:5px;">ICE Preliminary Open Interest — Futures</div>
    <div style="font-size:15px;color:{MUTED};margin-top:4px;">ICE Futures U.S. — preliminary, futures only</div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:22px;font-weight:700;color:{WHITE};">{report_date}</div>
    <div style="font-size:13px;color:{MUTED};margin-top:4px;">PRELIMINARY vs OFFICIAL</div>
  </div>
</div>

<div style="background:{ALT};padding:10px 36px;border-bottom:1px solid {BORDER};">
  <span style="font-size:14px;color:{GRAY};">
    Open interest shown is <b>ICE PRELIMINARY</b> for the {report_date} session.
    Changes measure against <b>OFFICIAL</b> open interest —
    DoD {b['DoD'] or '—'} · WoW {b['WoW'] or '—'} · MoM {b['MoM'] or '—'}.
  </span>
</div>

<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;padding:14px 20px;
            align-items:start;">
{blocks}
</div>
{note}
<div style="background:{ALT};padding:14px 36px;border-top:1px solid {BORDER};
            display:flex;justify-content:space-between;align-items:center;">
  <span style="font-size:11px;font-weight:700;color:{LGRAY};letter-spacing:1.5px;
               text-transform:uppercase;">VLM Commodities Ltd &nbsp;·&nbsp; Softs Desk</span>
  <span style="font-size:11px;font-weight:600;color:{MUTED};">vlmdata.com &nbsp;·&nbsp; {datetime.now():%Y-%m-%d %H:%M}</span>
</div>
</body></html>"""


def render_png(html, png_path):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print('WARN: playwright not installed — skipping PNG.')
        return None
    _CLIP = """() => { let b=0; for (const el of document.body.children) {
        const r = el.getBoundingClientRect(); if (r.bottom > b) b = r.bottom; }
        return Math.ceil(b) + 8; }"""
    with sync_playwright() as p:
        br = p.chromium.launch()
        pg = br.new_context(device_scale_factor=2).new_page()
        # Generous viewport: the clip-to-content evaluation below trims to the
        # real height, but a viewport shorter than the content silently CLIPS it
        # (a 2000px viewport truncated the sugar block at exactly 2000px).
        pg.set_viewport_size({'width': 1120, 'height': 4000})
        pg.set_content(html, wait_until='networkidle')
        h = pg.evaluate(_CLIP)
        w = pg.evaluate("() => Math.ceil(document.documentElement.scrollWidth)")
        pg.screenshot(path=str(png_path), clip={'x':0,'y':0,'width':w,'height':h})
        br.close()
    return str(png_path)


def write_xlsx(path, report_date, built, baselines):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('WARN: openpyxl not installed — skipping xlsx.')
        return None
    wb = Workbook(); ws = wb.active; ws.title = 'Prelim OI'
    hdr_fill = PatternFill('solid', fgColor='0F2744')
    hdr_font = Font(bold=True, color='FFFFFF')

    ws.append([f'ICE PRELIMINARY OPEN INTEREST — session {report_date}'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Preliminary values vs OFFICIAL baselines — '
               f"DoD {baselines['DoD']} · WoW {baselines['WoW']} · MoM {baselines['MoM']}"])
    ws.append([])

    for comm in COMMODITIES:
        d = built.get(comm)
        if not d:
            continue
        ws.append([COMM_NAMES[comm]])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        # Columns match the PNG exactly. ICE's own change figure is omitted here
        # too: it equals the computed DoD on every row, and carrying a column in
        # one artifact but not the other invites confusion later.
        head = ['Month', 'Prelim OI', 'DoD', 'WoW', 'MoM']
        ws.append(head)
        for c in range(1, len(head)+1):
            cell = ws.cell(ws.max_row, c)
            cell.fill, cell.font = hdr_fill, hdr_font
            cell.alignment = Alignment(horizontal='center')
        for r in d['rows']:
            ws.append([r['month'], r['oi'], r['chg']['DoD'], r['chg']['WoW'],
                       r['chg']['MoM']])
        t = d['total_chg']
        star = lambda lab: '*' if d['partial'][lab] else ''
        ws.append(['TOTAL', d['total_oi'],
                   f"{t['DoD']}{star('DoD')}" if star('DoD') else t['DoD'],
                   f"{t['WoW']}{star('WoW')}" if star('WoW') else t['WoW'],
                   f"{t['MoM']}{star('MoM')}" if star('MoM') else t['MoM']])
        for c in range(1, 6):
            ws.cell(ws.max_row, c).font = Font(bold=True)
        ws.append([])

    ws.append(['* change total sums only contracts with an official baseline; '
               'deep back months not carried by the generic chain are excluded '
               '(their OI is still in Prelim OI).'])
    for col, w in zip('ABCDE', (14, 14, 12, 12, 12)):
        ws.column_dimensions[col].width = w
    wb.save(path)
    return str(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--csv', help='explicit prelim CSV path')
    ap.add_argument('--no-png', action='store_true')
    ap.add_argument('--no-xlsx', action='store_true')
    ap.add_argument('--max-age-days', type=int, default=4,
                    help='refuse a CSV whose Report Date is older than this '
                         '(0 disables the check)')
    args = ap.parse_args()

    src = find_prelim_csv(args.csv)
    report_date, prelim, ice_chg = load_prelim(src)

    # Stale-CSV guard. The build is correct for whatever Report Date it is given,
    # which is the danger: emailing an OLD download produces a perfectly valid
    # report for the wrong session. Refuse rather than send right numbers for the
    # wrong day. 4 days covers a Friday session picked up the following Monday
    # (and a long weekend); --max-age-days 0 disables it for deliberate rebuilds.
    if args.max_age_days:
        try:
            age = (date.today() - datetime.strptime(report_date, '%Y-%m-%d').date()).days
        except ValueError:
            sys.exit(f'ERROR: unparseable Report Date {report_date!r}')
        if age > args.max_age_days:
            sys.exit(f'ERROR: prelim CSV is stale — Report Date {report_date} is '
                     f'{age} days old (limit {args.max_age_days}). This is almost '
                     f'certainly an older download. Re-download from ICE, or pass '
                     f'--max-age-days 0 to build it anyway.')
        if age < 0:
            sys.exit(f'ERROR: Report Date {report_date} is in the future — '
                     f'refusing to build.')
    sessions, by_date = load_official()
    baselines = pick_baselines(sessions, report_date)

    print(f'Source      : {src.name}')
    print(f'Session     : {report_date}  (PRELIMINARY)')
    for lab, n in WINDOWS:
        print(f'{lab} baseline: {baselines[lab] or "MISSING"}  (T-{n} official)')
    if report_date in sessions:
        print('NOTE: official data for this session already exists; baselines '
              'still use prior sessions so changes measure flow, not revision.')

    built = build_rows(prelim, ice_chg, by_date, baselines)

    # Reconciliation: our computed DoD should equal ICE's own change column
    # wherever a T-1 official baseline exists. A mismatch means the contract
    # mapping is wrong -- surface it rather than publish a silent error.
    agree = mismatch = 0
    for comm, d in built.items():
        for r in d['rows']:
            if r['chg']['DoD'] is None or r['ice_chg'] is None:
                continue
            if r['chg']['DoD'] == r['ice_chg']:
                agree += 1
            else:
                mismatch += 1
                print(f"  MISMATCH {comm} {r['month']}: "
                      f"computed {r['chg']['DoD']} vs ICE {r['ice_chg']}")
    print(f'Reconciled  : {agree} contracts agree with ICE, {mismatch} mismatch')

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = f'prelim_oi_{report_date}'
    html = build_html(report_date, built, baselines, src.name)
    (OUT_DIR / f'{stem}.html').write_text(html, encoding='utf-8')

    if not args.no_png:
        p = render_png(html, OUT_DIR / f'{stem}.png')
        if p: print(f'PNG         : {p}')
    if not args.no_xlsx:
        x = write_xlsx(OUT_DIR / f'{stem}.xlsx', report_date, built, baselines)
        if x: print(f'XLSX        : {x}')


if __name__ == '__main__':
    main()
